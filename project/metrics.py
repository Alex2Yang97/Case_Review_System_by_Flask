import openpyxl
import pandas as pd

from .utils import connect_db, SQL_TYPE


def get_month_year(m, y, str_type="full"):
    if str_type == "full":
        result = str(y)
        months = {
          1: "January", 2: "February", 3: "March", 4: "April", 5: "May", 6: "June",
          7: "July", 8: "August", 9: "September", 10: "October", 11: "November", 12: "December",
          }
        result = months[m] + " " + result
    else:
        result = str(y)[-2:]
        months = {
          1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
          7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
          }
        result = months[m] + "-" + result
    return result


def annual_chats_for_metrics(year, case_type, sql_type=SQL_TYPE):
    cursor, cnxn = connect_db()

    allcase_df = pd.DataFrame()
    for i in range(1, 5):
        
        if sql_type == "sql server":
            allcase_sql = f'''
                SELECT Risk_Rating, COUNT(Risk_Rating) AS cnt 
                    FROM CaseTracking_local 
                    WHERE (Month(Scheduled_Due_Date) BETWEEN {i*3-2} AND {i*3}) AND 
                    Year(Scheduled_Due_Date) = {year} AND Case_Status <> 'Removed' 
                    AND (Case_Type IS NULL OR {case_type}) 
                    GROUP BY Risk_Rating '''
        else:
            allcase_sql = f'''
                SELECT Risk_Rating, COUNT(Risk_Rating) AS cnt 
                    FROM CaseTracking_local 
                    WHERE ( ltrim(strftime('%m', Scheduled_Due_Date), "0") BETWEEN {i*3-2} AND {i*3}) AND 
                    strftime('%Y', Scheduled_Due_Date) = {year} AND Case_Status <> 'Removed' 
                    AND (Case_Type IS NULL OR {case_type}) 
                    GROUP BY Risk_Rating '''
        
        allcase_i = pd.read_sql(allcase_sql, cnxn)

        if not allcase_i.empty:
            allcase_i["Due Date"] = f"{i}Q{year}"
            allcase_i = allcase_i.pivot(index='Due Date', columns='Risk_Rating', values='cnt').reset_index(drop=False)
        else:
            allcase_i = pd.DataFrame({
                "Due Date": [f"{i}Q{year}"],
                "High": [0], "Medium": [0], "Low": [0]
            })

        allcase_df = allcase_df.append(allcase_i)
        
    cursor.close()
    cnxn.close()
    
    allcase_df = allcase_df.fillna(0).sort_values(by="Due Date")
    allcase_df["Total"] = allcase_df["High"] + allcase_df["Medium"] + allcase_df["Low"]
    return allcase_df


def tables_for_metrics(month, year, case_status, case_type, sql_type=SQL_TYPE): # Closed
    res_df = pd.DataFrame({"Risk_Rating": ["High", "Medium", "Low"]})
    month_year_str_lst = []
    
    if month > 1 and month < 12:
        months_lst = [month-1, month, month+1]
    elif month == 1:
        months_lst = [12, 1, 2]
    else:
        months_lst = [11, 12, 1]
        
    for m in months_lst:
        month_year_str = get_month_year(m, year)
        month_year_str_lst.append(month_year_str)

        cursor, cnxn = connect_db()
        if sql_type == "sql server":
            metrics_sql = f'''SELECT Risk_Rating, COUNT(Risk_Rating) AS cnt FROM CaseTracking_local 
                WHERE Month(Scheduled_Due_Date) = {m} and Year(Scheduled_Due_Date) = {year} AND 
                {case_status} AND (Case_Type IS NULL OR {case_type}) 
                GROUP BY Risk_Rating'''
        else:
            metrics_sql = f'''
                SELECT Risk_Rating, COUNT(Risk_Rating) AS cnt FROM CaseTracking_local 
                WHERE ltrim(strftime('%m', Scheduled_Due_Date), "0") = {m} and strftime('%Y', Scheduled_Due_Date) = {year} AND 
                {case_status} AND (Case_Type IS NULL OR {case_type}) 
                GROUP BY Risk_Rating'''
        
        month_df = pd.read_sql(metrics_sql, cnxn)
        cursor.close()
        cnxn.close()
        
        if month_df.empty:
            month_df = pd.DataFrame({
                "Risk_Rating": ["High", "Medium", "Low"],
                month_year_str: [0, 0, 0], })
        else:
            month_df = month_df.rename(columns={"cnt": month_year_str})

        res_df = pd.merge(res_df, month_df, on="Risk_Rating", how="left")
        res_df = res_df.fillna(0)
    return res_df, month_year_str_lst


def check_cases_prop(all_cnt, close_cnt):
    if all_cnt == 0:
        return "N/A", "FFFFFF"
    elif all_cnt == close_cnt:
        return "Completed", "ADD8E6"
    else:
        return "On target", "90EE90"


def get_recom_metrics(last_month, ytd_year, ytd_months_lst, sql_type=SQL_TYPE):
      
    cursor, cnxn = connect_db()
    
    if sql_type == "sql server":
        sql_recom_all_months = f'''
            SELECT Scheduled_Due_Date, Risk_Rating as RR, Recommendation_local.Status as Status 
            FROM CaseTracking_local RIGHT JOIN Recommendation_local ON CaseTracking_local.Case_ID = Recommendation_local.Case_ID 
            WHERE Recommendation_local.Case_ID IS NOT NULL and 
            (Year(Scheduled_Due_Date) = {ytd_year} or Year(Scheduled_Due_Date) = {ytd_year-1})'''
    else:
        sql_recom_all_months = f'''
            SELECT Scheduled_Due_Date, Risk_Rating as RR, Recommendation_local.Status as Status 
            FROM Recommendation_local LEFT JOIN CaseTracking_local ON CaseTracking_local.Case_ID = Recommendation_local.Case_ID 
            WHERE Recommendation_local.Case_ID IS NOT NULL and 
            (strftime('%Y', Scheduled_Due_Date) = {ytd_year} or strftime('%Y', Scheduled_Due_Date) = {ytd_year-1}) '''
        
    recom_res = pd.read_sql(sql_recom_all_months, cnxn)
    cursor.close()
    cnxn.close()
    
    recom_res["table_date"] = pd.to_datetime(recom_res["Scheduled_Due_Date"])
    recom_res["due_month"] = recom_res["table_date"].dt.month
    recom_res["due_year"] = recom_res["table_date"].dt.year

    recom_ytd = recom_res[
        (recom_res["due_year"]==ytd_year) & (recom_res["due_month"].isin(ytd_months_lst)) & 
        ((recom_res["Status"]=="Closed") | (recom_res["Status"]=="Open"))]
    grouped_recom_ytd = recom_ytd.groupby(["Status", "RR", "due_month"])[["table_date"]].count().reset_index().rename(columns={"table_date": "total"})

    last_month_recom = grouped_recom_ytd[grouped_recom_ytd["due_month"]==last_month]

    ytd_recom = grouped_recom_ytd.groupby(["Status", "RR"])["total"].sum().reset_index()

    ytd_months = grouped_recom_ytd[(grouped_recom_ytd["Status"]=="Closed") | (grouped_recom_ytd["Status"]=="Open")]
    ytd_months = ytd_months.groupby(["due_month", "RR"])["total"].sum().reset_index()

    return last_month_recom, ytd_recom, ytd_months


def get_sar_metrics(last_month, ytd_year, ytd_months_lst, sql_type=SQL_TYPE):
    cursor, cnxn = connect_db()
    if sql_type == "sql server":
        sql_sar_all_months = f'''
            SELECT Scheduled_Due_Date, Risk_Rating as RR, Referral_Necessary 
            FROM CaseTracking_local RIGHT JOIN SARref_local ON CaseTracking_local.Case_ID = SARref_local.Case_ID 
            WHERE SARref_local.Case_ID IS NOT NULL and 
            (Year(Scheduled_Due_Date) = {ytd_year} or Year(Scheduled_Due_Date) = {ytd_year-1}) '''
    else:
        sql_sar_all_months = f'''
            SELECT Scheduled_Due_Date, Risk_Rating as RR, Referral_Necessary 
            FROM SARref_local LEFT JOIN CaseTracking_local ON CaseTracking_local.Case_ID = SARref_local.Case_ID 
            WHERE SARref_local.Case_ID IS NOT NULL and 
            (strftime('%Y', Scheduled_Due_Date) = {ytd_year} or strftime('%Y', Scheduled_Due_Date) = {ytd_year-1}) '''
        
    sar_res = pd.read_sql(sql_sar_all_months, cnxn)
    cursor.close()
    cnxn.close()
    
    if sql_type == "sql server":
        sar_res["table_date"] = sar_res["Scheduled_Due_Date"] + pd.DateOffset(1)
    else:
        sar_res["table_date"] = pd.to_datetime(sar_res["Scheduled_Due_Date"]) + pd.DateOffset(1)
        
    sar_res["due_month"] = sar_res["table_date"].dt.month
    sar_res["due_year"] = sar_res["table_date"].dt.year

    sar_ytd = sar_res[
        (sar_res["due_year"]==ytd_year) & (sar_res["due_month"].isin(ytd_months_lst))]
    grouped_sar_ytd = sar_ytd.groupby(["Referral_Necessary", "RR", "due_month"])[["table_date"]].count().reset_index().rename(columns={"table_date": "total"})

    last_month_sar = grouped_sar_ytd[grouped_sar_ytd["due_month"]==last_month]
    ytd_sar = grouped_sar_ytd.groupby(["Referral_Necessary", "RR"])["total"].sum().reset_index()
    ytd_months = grouped_sar_ytd[grouped_sar_ytd["Referral_Necessary"]=="Y"]

    return last_month_sar, ytd_sar, ytd_months


def get_value_for_month(m, ytd_months):
    month_high = ytd_months[(ytd_months["due_month"]==m) & (ytd_months["RR"]=="High")]["total"]
    month_high = 0 if month_high.empty else month_high.item()

    month_med = ytd_months[(ytd_months["due_month"]==m) & (ytd_months["RR"]=="Medium")]["total"]
    month_med = 0 if month_med.empty else month_med.item()

    month_low = ytd_months[(ytd_months["due_month"]==m) & (ytd_months["RR"]=="Low")]["total"]
    month_low = 0 if month_low.empty else month_low.item()
    
    month_total = month_high + month_med + month_low
    
    return month_high, month_med, month_low, month_total


def recom_value_for_rr(rr, recom_res):
    recom_open = recom_res[(recom_res["Status"]=="Open") & (recom_res["RR"]==rr)]["total"]
    recom_open = 0 if recom_open.empty else recom_open.item()
    
    recom_closed = recom_res[(recom_res["Status"]=="Closed") & (recom_res["RR"]==rr)]["total"]
    recom_closed = 0 if recom_closed.empty else recom_closed.item()
    
    recom_total = recom_open + recom_closed
    
    return recom_open, recom_closed, recom_total


def sar_value_for_rr(rr, sar_res):
    sar_not_required = sar_res[(sar_res["Referral_Necessary"]=="N") & (sar_res["RR"]==rr)]["total"]
    sar_not_required = 0 if sar_not_required.empty else sar_not_required.item()
    
    sar_made = sar_res[(sar_res["Referral_Necessary"]=="Y") & (sar_res["RR"]==rr)]["total"]
    sar_made = 0 if sar_made.empty else sar_made.item()
    
    sar_total = sar_not_required + sar_made
    
    return sar_not_required, sar_made, sar_total
  
  
def gen_case_metrics(year, month, filename):
    wb = openpyxl.load_workbook(filename)
    sheet1 = wb["Case Metrics"]

    annual_case_metrics = annual_chats_for_metrics(year, case_type="Case_Type <> 'RMB'")

    columns = ["X", "Y", "Z", "AA", "AB"]
    names = ["Due Date", "High", "Medium", "Low"]
    for col, name in zip(columns, names):
        cells = [sheet1[f"{col}5"], sheet1[f"{col}6"], sheet1[f"{col}7"], sheet1[f"{col}8"]]
        numbers = annual_case_metrics[f"{name}"].tolist()
        for cell, number in zip(cells, numbers):
            cell.value = number

    edd_all, month_year_str_lst = tables_for_metrics(
        month, year, case_status="Case_Status <> 'Removed'", case_type="Case_Type <> 'RMB'")

    edd_close, _ = tables_for_metrics(
        month, year, case_status="Case_Status = 'Closed'", case_type="Case_Type <> 'RMB'")

    case_columns = ["B", "D", "F"]
    status_columns = ["C", "E", "G"]
    for i, col in enumerate(case_columns):
        
        due_date = month_year_str_lst[i]
        status_col = status_columns[i]
        
        if i == 0:
            sheet1["A1"] = f"FCB EDD RMB Clearing Bank Monthly Status ({due_date} with 3-Month Outlook)"
            sheet1["A9"] = f"* The {due_date} EDD Case Population does not include 4 monthly RMB Clearing Bank customers review."
        
        sheet1[f"{col}3"].value = due_date
        sheet1[f"{col}4"].value = "Case Due: " + due_date
        
        # High
        all_cnt = edd_all[edd_all["Risk_Rating"]=="High"][due_date].item()
        close_cnt = edd_close[edd_close["Risk_Rating"]=="High"][due_date].item()
        status, color = check_cases_prop(all_cnt, close_cnt)
        sheet1[f"{col}5"].value = all_cnt
        sheet1[f"{status_col}5"].value = status
        fill = openpyxl.styles.PatternFill(
            start_color=color,
            end_color=color,
            fill_type="solid",
        )
        sheet1[f"{status_col}5"].fill = fill
        
        # Medium
        all_cnt = edd_all[edd_all["Risk_Rating"]=="Medium"][due_date].item()
        close_cnt = edd_close[edd_close["Risk_Rating"]=="Medium"][due_date].item()
        status, color = check_cases_prop(all_cnt, close_cnt)
        sheet1[f"{col}6"].value = all_cnt
        sheet1[f"{status_col}6"].value = status
        fill = openpyxl.styles.PatternFill(
            start_color=color,
            end_color=color,
            fill_type="solid",
        )
        sheet1[f"{status_col}6"].fill = fill
        
        # Low
        all_cnt = edd_all[edd_all["Risk_Rating"]=="Low"][due_date].item()
        close_cnt = edd_close[edd_close["Risk_Rating"]=="Low"][due_date].item()
        status, color = check_cases_prop(all_cnt, close_cnt)
        sheet1[f"{col}7"].value = all_cnt
        sheet1[f"{status_col}7"].value = status
        fill = openpyxl.styles.PatternFill(
            start_color=color,
            end_color=color,
            fill_type="solid",
        )
        sheet1[f"{status_col}7"].fill = fill
        
        sheet1[f"{col}8"].value = edd_all[due_date].sum()

    wb.save(filename)
    return
  
  
def gen_rmb_case_metrics(year, month, filename):
    wb = openpyxl.load_workbook(filename)
    sheet2 = wb["RMB Case Metrics"]

    annual_rmb_metrics = annual_chats_for_metrics(year, case_type="Case_Type = 'RMB'")

    columns = ["X", "Y", "Z", "AA", "AB"]
    names = ["Due Date", "High", "Medium", "Low"]
    for col, name in zip(columns, names):
        cells = [sheet2[f"{col}5"], sheet2[f"{col}6"], sheet2[f"{col}7"], sheet2[f"{col}8"]]
        numbers = annual_rmb_metrics[f"{name}"].tolist()
        for cell, number in zip(cells, numbers):
            cell.value = number

    rmb_all, month_year_str_lst = tables_for_metrics(
        month, year, case_status="Case_Status <> 'Removed'", case_type="Case_Type = 'RMB'")

    rmb_close, _ = tables_for_metrics(
        month, year, case_status="Case_Status = 'Closed'", case_type="Case_Type = 'RMB'")

    case_columns = ["B", "D", "F"]
    status_columns = ["C", "E", "G"]
    for i, col in enumerate(case_columns):
        
        due_date = month_year_str_lst[i]
        status_col = status_columns[i]
        
        if i == 0:
            sheet2["A1"] = f"FCB EDD RMB Clearing Bank Monthly Status ({due_date} with 3-Month Outlook)"
            
        sheet2[f"{col}3"].value = due_date
        sheet2[f"{col}4"].value = "Case Due: " + due_date
        
        # High
        all_cnt = rmb_all[rmb_all["Risk_Rating"]=="High"][due_date].item()
        close_cnt = rmb_close[rmb_close["Risk_Rating"]=="High"][due_date].item()
        status, color = check_cases_prop(all_cnt, close_cnt)
        sheet2[f"{col}5"].value = all_cnt
        sheet2[f"{status_col}5"].value = status
        fill = openpyxl.styles.PatternFill(
            start_color=color,
            end_color=color,
            fill_type="solid",
        )
        sheet2[f"{status_col}5"].fill = fill
        
        # Medium
        all_cnt = rmb_all[rmb_all["Risk_Rating"]=="Medium"][due_date].item()
        close_cnt = rmb_close[rmb_close["Risk_Rating"]=="Medium"][due_date].item()
        status, color = check_cases_prop(all_cnt, close_cnt)
        sheet2[f"{col}6"].value = all_cnt
        sheet2[f"{status_col}6"].value = status
        fill = openpyxl.styles.PatternFill(
            start_color=color,
            end_color=color,
            fill_type="solid",
        )
        sheet2[f"{status_col}6"].fill = fill
        
        # Low
        all_cnt = rmb_all[rmb_all["Risk_Rating"]=="Low"][due_date].item()
        close_cnt = rmb_close[rmb_close["Risk_Rating"]=="Low"][due_date].item()
        status, color = check_cases_prop(all_cnt, close_cnt)
        sheet2[f"{col}7"].value = all_cnt
        sheet2[f"{status_col}7"].value = status
        fill = openpyxl.styles.PatternFill(
            start_color=color,
            end_color=color,
            fill_type="solid",
        )
        sheet2[f"{status_col}7"].fill = fill
        
        sheet2[f"{col}8"].value = rmb_all[due_date].sum()

    wb.save(filename)
    return


def gen_recom_metrics(year, month, filename):
    wb = openpyxl.load_workbook(filename)
    sheet3 = wb["Recommendation"]

    last_month = month - 1 if month != 1 else 12
    ytd_year = year if last_month != 12 else year - 1
    ytd_months_lst = list(range(1, last_month+1))

    last_month_str = get_month_year(last_month, ytd_year, str_type="abb")
    sheet3["B3"].value = last_month_str
    sheet3["F3"].value = str(ytd_year) + " YTD"

    last_month_recom, ytd_recom, ytd_months = get_recom_metrics(last_month, ytd_year, ytd_months_lst)

    rr_lst = ["High", "Medium", "Low"]

    for i, rr in enumerate(rr_lst):
        recom_open_month, recom_closed_month, recom_total_month = recom_value_for_rr(rr, last_month_recom)
        sheet3[f"B{i+5}"].value = recom_total_month
        sheet3[f"C{i+5}"].value = recom_closed_month
        sheet3[f"D{i+5}"].value = recom_open_month
        sheet3[f"E{i+5}"].value = 0
        
        recom_open_year, recom_closed_year, recom_total_year = recom_value_for_rr(rr, ytd_recom)
        sheet3[f"F{i+5}"].value = recom_total_year
        sheet3[f"G{i+5}"].value = recom_closed_year
        sheet3[f"H{i+5}"].value = recom_open_year
        sheet3[f"I{i+5}"].value = 0

    for m in range(1, 13):
        date_str = get_month_year(m, ytd_year, str_type="abb")
        month_high, month_med, month_low, month_total = get_value_for_month(m, ytd_months)
        
        sheet3[f"AA{m+2}"].value = date_str
        sheet3[f"AB{m+2}"].value = month_high
        sheet3[f"AC{m+2}"].value = month_med
        sheet3[f"AD{m+2}"].value = month_low
        sheet3[f"AE{m+2}"].value = month_total
        
    wb.save(filename)
    return
  
  
def gen_sar_metrics(year, month, filename):
    wb = openpyxl.load_workbook(filename)
    sheet4 = wb["Referral"]

    last_month = month - 1 if month != 1 else 12
    ytd_year = year if last_month != 12 else year - 1
    ytd_months_lst = list(range(1, last_month+1))

    last_month_str = get_month_year(last_month, ytd_year, str_type="abb")
    sheet4["B3"].value = last_month_str
    sheet4["F3"].value = str(ytd_year) + " YTD"

    last_month_sar, ytd_sar, ytd_months = get_sar_metrics(last_month, ytd_year, ytd_months_lst)

    rr_lst = ["High", "Medium", "Low"]

    for i, rr in enumerate(rr_lst):
        sar_not_required_month, sar_made_month, sar_total_month = sar_value_for_rr(rr, last_month_sar)
        sheet4[f"B{i+5}"].value = sar_total_month
        sheet4[f"C{i+5}"].value = sar_not_required_month
        sheet4[f"D{i+5}"].value = sar_made_month
        sheet4[f"E{i+5}"].value = 0
        
        sar_not_required_year, sar_made_year, sar_total_year = sar_value_for_rr(rr, ytd_sar)
        sheet4[f"F{i+5}"].value = sar_total_year
        sheet4[f"G{i+5}"].value = sar_not_required_year
        sheet4[f"H{i+5}"].value = sar_made_year
        sheet4[f"I{i+5}"].value = 0

    for m in range(1, 13):
        date_str = get_month_year(m, ytd_year, str_type="abb")
        month_high, month_med, month_low, month_total = get_value_for_month(m, ytd_months)
        
        sheet4[f"AA{m+2}"].value = date_str
        sheet4[f"AB{m+2}"].value = month_high
        sheet4[f"AC{m+2}"].value = month_med
        sheet4[f"AD{m+2}"].value = month_low
        sheet4[f"AE{m+2}"].value = month_total

    wb.save(filename)
    return

